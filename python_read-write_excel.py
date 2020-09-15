# 导包
import openpyxl
from openpyxl.styles import Font, colors, Alignment

# 读取Excel文件到内存中并生成wb对象
wb = openpyxl.load_workbook('File/123.xlsx')

# 获取Excel所有sheet名称
sheetNames = wb.sheetnames
print(sheetNames)

# 根据sheet名称获取sheet对象
sheet = wb['SheetJS']
print(sheet.title)

# 获取当前正在显示的sheet
sheet = wb.active
print(sheet.title)

# 用单元格下标获取a2单元格的值
a2 = sheet['A2']
print(a2.value)
# column列；row行
print(a2.column, a2.row)

# 用cell函数行与列获取单元格d2单元格的值
d2 = sheet.cell(row=2, column=4)
print(d2.value)

# 获取最大行
max_row = sheet.max_row
print(max_row)
# 获取最大列
max_column = sheet.max_column
print(max_column)

# 遍历所有行的数据
# sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# 遍历所有列的数据
# sheet.columns类似，不过里面是每个tuple是每一列的单元格
for column in sheet.columns:
    for cell in column:
        print(cell.value)

# 获取某一行的数据
rowList = list(sheet.rows)
for cell in rowList[1]:
    print(cell.value)

# 获取某一列的数据
columnList = list(sheet.columns)
for cell in columnList[3]:
    print(cell.value)

# 获得任意区间的单元格
maxRow = sheet.max_row
maxColumn = sheet.max_column
for i in range(sheet.min_row, maxRow+1):
    for j in range(sheet.min_column, maxColumn+1):
        print(sheet.cell(row=i, column=j).value)

# 根据下标获取任意区间内的单元格
for rowCell in sheet['B1':'F2']:
    for cell in rowCell:
        print(cell.value)

# 根据列的数字返回字母
print(openpyxl.utils.get_column_letter(3))
# 根据字母返回列的数字
print(openpyxl.utils.column_index_from_string('D'))

# 写入数据到Excel表
# 创建一个未保存的工作表
wb = openpyxl.Workbook()
print(wb.sheetnames)    # 默认的sheet名字为Sheet
# 直接赋值可以修改sheet的名称
wb['Sheet'].title = '写入数据'
print(wb.sheetnames)
wb.create_sheet('写入数据2', index=1)
print(wb.sheetnames)
sheet = wb['写入数据']
# 删除某个工作表
wb.remove(sheet)
del wb['写入数据2']
print(wb.sheetnames)

# 写入单元格
sheet = wb.create_sheet('写入数据3')
sheet['A1'] = '测试写入数据到A1单元格'
print(sheet['A1'].value)
# 写入公式
sheet['B1'] = '=AVERAGE(B2:B8)'
# 读取的时候需要加上data_only=True读到返回的是数字,不加参数返回公式本身
print(sheet['B1'].value)
print(sheet.cell(row=2, column=1).value)

# 一次添加多行数据 append()函数，从第一行空白行 一行一行的写入
# 添加一行
row = [1, 2, 3, 4, 5, 6]
sheet.append(row)
for cell in list(sheet.rows)[2]:
    print(cell.value)

# 添加多行
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
for row in rows:
    sheet.append(row)
for cell in list(sheet.rows)[4]:
    print(cell.value)

# 按列写入 zip()函数 最后的元组个数是由原来每个参数（可迭代对象）的最短长度决定的
columnsList = list(zip(*rows))
print(columnsList)
for column in columnsList:
    sheet.append(column)
for cell in list(sheet.rows)[6]:
    print(cell.value)

# 保存文件
wb.save('File/写入数据.xlsx')

# 设置单元格风格——style
# 设置字体
bold_itatic_24_font = Font(
    name='等线', size=24, color=colors.RED, italic=True, bold=True)
sheet.cell(row=5, column=1).font = bold_itatic_24_font
wb.save('File/写入数据.xlsx')
# 设置对齐方式
# 除了center，还可以使用right、left等等参数
sheet['A5'].alignment = Alignment(horizontal='center', vertical='center')
wb.save('File/写入数据.xlsx')

# 设置行高和列宽
sheet.row_dimensions[2].height = 50
sheet.column_dimensions['C'].width = 60
wb.save('File/写入数据.xlsx')

# 合并和拆分单元格
sheet.merge_cells('B1:F1')  # 合并一行中的几个单元格
sheet.merge_cells('B2:D3')  # 合并一个矩形区域中的单元格
sheet.unmerge_cells('B1:F1')# 拆分单元格
wb.save('File/写入数据.xlsx')